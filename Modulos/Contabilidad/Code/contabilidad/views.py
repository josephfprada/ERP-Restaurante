from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, F
from django.contrib import messages
from datetime import date, timedelta
from .models import ComprobanteMaestro, Factura, TipoCuenta, CuentaMayor, SubCuenta, DetalleComprobante, Comprobante
from .forms import FacturaForm, TipoCuentaForm, CuentaMayorForm, SubCuentaForm, ComprobanteForm
from django.http import HttpResponse
from django.utils.text import slugify
from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font, Alignment, Border, Side # type: ignore



# Create your views here.
def login(request):
    return render(request, 'contabilidad/login.html')

def index(request):

    hoy = date.today()
    hace_7_dias = hoy - timedelta(days=7)
    hace_30_dias = hoy - timedelta(days=30)

    comprobantes_recientes = ComprobanteMaestro.objects.filter(
        fecha_comprobante__gte=hace_30_dias,
        estado='CONTABILIZADO' 
    )
    total_debitos_recientes = comprobantes_recientes.aggregate(Sum('total_debito'))['total_debito__sum'] or 0
    total_creditos_recientes = comprobantes_recientes.aggregate(Sum('total_credito'))['total_credito__sum'] or 0
    conteo_comprobantes_recientes = comprobantes_recientes.count()

    facturas_emitidas = Factura.objects.filter(estado='EMITIDA').count()
    facturas_pagadas = Factura.objects.filter(estado='PAGADA').count()
    facturas_pendientes_pago = Factura.objects.filter(estado='EMITIDA').exclude(estado='PAGADA').count()


    context = {
        'total_debitos_recientes': total_debitos_recientes,
        'total_creditos_recientes': total_creditos_recientes,
        'conteo_comprobantes_recientes': conteo_comprobantes_recientes,
        'periodo_comprobantes': "últimos 30 días",
        'facturas_emitidas': facturas_emitidas,
        'facturas_pagadas': facturas_pagadas,
        'facturas_pendientes_pago': facturas_pendientes_pago,
        'fecha_actual': hoy,
    }

    return render(request, 'contabilidad/index.html', context)


def gestion_cuentas_contables(request):

    tipo_cuenta_form = TipoCuentaForm()
    cuenta_mayor_form = CuentaMayorForm()
    sub_cuenta_form = SubCuentaForm()

    if request.method == 'POST':
        if 'submit_tipo_cuenta' in request.POST:
            tipo_cuenta_form = TipoCuentaForm(request.POST)
            if tipo_cuenta_form.is_valid():
                tipo_cuenta_form.save()
                messages.success(request, 'Tipo de Cuenta creado exitosamente.')
                return redirect('CuentasContables')
            else:
                messages.error(request, 'Error al crear Tipo de Cuenta. Revisa los datos.')

        elif 'submit_cuenta_mayor' in request.POST:
            cuenta_mayor_form = CuentaMayorForm(request.POST)
            if cuenta_mayor_form.is_valid():
                cuenta_mayor_form.save()
                messages.success(request, 'Cuenta Mayor creada exitosamente.')
                return redirect('CuentasContables')
            else:
                messages.error(request, 'Error al crear Cuenta Mayor. Revisa los datos.')

        elif 'submit_sub_cuenta' in request.POST:
            sub_cuenta_form = SubCuentaForm(request.POST)
            if sub_cuenta_form.is_valid():
                sub_cuenta_form.save()
                messages.success(request, 'Sub-Cuenta creada exitosamente.')
                return redirect('CuentasContables')
            else:
                messages.error(request, 'Error al crear Sub-Cuenta. Revisa los datos.')

    tipos_cuenta = TipoCuenta.objects.all().order_by('nombre')
    cuentas_mayores = CuentaMayor.objects.all().order_by('id')
    sub_cuentas = SubCuenta.objects.all().order_by('id')

    context = {
        'tipo_cuenta_form': tipo_cuenta_form,
        'cuenta_mayor_form': cuenta_mayor_form,
        'sub_cuenta_form': sub_cuenta_form,
        'tipos_cuenta': tipos_cuenta,
        'cuentas_mayores': cuentas_mayores,
        'sub_cuentas': sub_cuentas,
    }
    return render(request, 'contabilidad/cuentas_contables.html', context)

def manage_accounts_view(request):

    if request.method == 'POST':
        if 'submit_tipo_cuenta' in request.POST:
            tipo_cuenta_form = TipoCuentaForm(request.POST)
            if tipo_cuenta_form.is_valid():
                tipo_cuenta_form.save()
                messages.success(request, 'Tipo de cuenta creado exitosamente!')
                return redirect('your_current_view_name') 
            else:
                messages.error(request, 'Error al crear el tipo de cuenta.')
        elif 'submit_cuenta_mayor' in request.POST:
            cuenta_mayor_form = CuentaMayorForm(request.POST)
            if cuenta_mayor_form.is_valid():
                cuenta_mayor_form.save()
                messages.success(request, 'Cuenta mayor creada exitosamente!')
                return redirect('your_current_view_name')
            else:
                messages.error(request, 'Error al crear la cuenta mayor.')
        elif 'submit_sub_cuenta' in request.POST:
            sub_cuenta_form = SubCuentaForm(request.POST)
            if sub_cuenta_form.is_valid():
                sub_cuenta_form.save()
                messages.success(request, 'Sub-cuenta creada exitosamente!')
                return redirect('your_current_view_name')
            else:
                messages.error(request, 'Error al crear la sub-cuenta.')

    else:
        tipo_cuenta_form = TipoCuentaForm()
        cuenta_mayor_form = CuentaMayorForm()
        sub_cuenta_form = SubCuentaForm()

    tipos_cuenta = TipoCuenta.objects.all()
    cuentas_mayores = CuentaMayor.objects.all()
    sub_cuentas = SubCuenta.objects.all()

    context = {
        'tipo_cuenta_form': tipo_cuenta_form,
        'cuenta_mayor_form': cuenta_mayor_form,
        'sub_cuenta_form': sub_cuenta_form,
        'tipos_cuenta': tipos_cuenta,
        'cuentas_mayores': cuentas_mayores,
        'sub_cuentas': sub_cuentas,
    }
    return render(request, 'your_template_name.html', context)


def delete_tipo_cuenta(request, pk):
    tipo_cuenta = get_object_or_404(TipoCuenta, pk=pk)

    if request.method == 'POST':
        tipo_cuenta.delete()
        messages.success(request, f'Tipo de cuenta "{tipo_cuenta.nombre}" eliminado exitosamente.')
        return redirect('CuentasContables')
    else:
        messages.error(request, 'Método no permitido para esta operación.')
        return redirect('CuentasContables')
    


def gestion_facturas(request):
    if request.method == 'POST':
        form = FacturaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Factura creada exitosamente!')
            return redirect('GestionFacturas')
        else:
            messages.error(request, 'Hubo un error al crear la factura. Por favor, revisa los datos.')
    else:
        form = FacturaForm()

    facturas = Factura.objects.all() 

    context = {
        'form': form,
        'facturas': facturas,
    }
    return render(request, 'contabilidad/gestion_facturas.html', context)


def ver_factura(request, pk):
    factura = get_object_or_404(Factura, pk=pk)
    context = {
        'factura': factura
    }
    return render(request, 'contabilidad/ver_facturas.html', context) 

def EditarFactura(request, pk):
    factura = get_object_or_404(Factura, pk=pk)

    if request.method == 'POST':
        form = FacturaForm(request.POST, instance=factura)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Factura actualizada exitosamente!')
            return redirect('GestionFacturas')
        else:
            messages.error(request, 'Hubo un error al actualizar la factura. Por favor, revisa los datos.')
    else:
        form = FacturaForm(instance=factura)

    context = {
        'form': form,
        'factura': factura,
        'editing': True
    }
    return render(request, 'contabilidad/gestion_facturas.html', context)

def eliminar_factura(request, pk):
    factura = get_object_or_404(Factura, pk=pk)
    if request.method == 'POST':
        factura.delete()
        messages.success(request, f'Factura #{factura.numero_factura} eliminada exitosamente.')
        return redirect('GestionFacturas')
    else:
        messages.error(request, 'Acceso inválido para eliminar factura.')
        return redirect('GestionFacturas')

def gestion_comprobantes(request):
    comprobantes = Comprobante.objects.all()
    form = ComprobanteForm() # Formulario para crear un nuevo comprobante si se desea desde aquí

    if request.method == 'POST':
        form = ComprobanteForm(request.POST, request.FILES) # Important: request.FILES for file uploads
        if form.is_valid():
            form.save()
            messages.success(request, '¡Comprobante subido exitosamente!')
            return redirect('GestionComprobantes')
        else:
            messages.error(request, 'Hubo un error al subir el comprobante. Por favor, revisa los datos.')

    context = {
        'comprobantes': comprobantes,
        'form': form,
    }
    return render(request, 'contabilidad/gestion_comprobantes.html', context)

def subir_comprobante_general(request):
    if request.method == 'POST':
        form = ComprobanteForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Comprobante subido exitosamente!')
            return redirect('GestionComprobantes') # Redirige a la lista general
        else:
            messages.error(request, 'Hubo un error al subir el comprobante. Por favor, revisa los datos.')
    else:
        form = ComprobanteForm()
    
    context = {
        'form': form,
        'title': 'Subir Nuevo Comprobante',
        'is_upload_page': True, # Flag para la plantilla si necesitas diferenciar
    }
    return render(request, 'contabilidad/subir_comprobante.html', context)


def subir_comprobante_para_factura(request, factura_pk):
    factura = get_object_or_404(Factura, pk=factura_pk)
    if request.method == 'POST':
        form = ComprobanteForm(request.POST, request.FILES)
        if form.is_valid():

            comprobante = form.save(commit=False)
            comprobante.factura = factura
            comprobante.save()
            messages.success(request, f'¡Comprobante para factura {factura.numero_factura} subido exitosamente!')
            return redirect('VerFacturas', pk=factura.pk) # Redirige a la vista de la factura
        else:
            messages.error(request, 'Hubo un error al subir el comprobante. Por favor, revisa los datos.')
    else:

        form = ComprobanteForm(initial={'factura': factura})
   
    context = {
        'form': form,
        'factura': factura,
        'title': f'Subir Comprobante para Factura #{factura.numero_factura}',
    }
    return render(request, 'contabilidad/subir_comprobante.html', context)


def editar_comprobante(request, pk):
    comprobante = get_object_or_404(Comprobante, pk=pk)
    if request.method == 'POST':
        form = ComprobanteForm(request.POST, request.FILES, instance=comprobante)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Comprobante actualizado exitosamente!')
            return redirect('GestionComprobantes') # Redirige a la lista general
        else:
            messages.error(request, 'Hubo un error al actualizar el comprobante. Por favor, revisa los datos.')
    else:
        form = ComprobanteForm(instance=comprobante)
    
    context = {
        'form': form,
        'comprobante': comprobante, # Para poder mostrar detalles del comprobante
        'title': f'Editar Comprobante para Factura #{comprobante.factura.numero_factura}',
    }
    return render(request, 'contabilidad/editar_comprobante.html', context) # Considera una plantilla dedicada


def eliminar_comprobante(request, pk):
    comprobante = get_object_or_404(Comprobante, pk=pk)
    if request.method == 'POST':
        comprobante.delete()
        messages.success(request, '¡Comprobante eliminado exitosamente!')
        return redirect('GestionComprobantes')
    
    context = {
        'comprobante': comprobante,
        'title': 'Confirmar Eliminación de Comprobante',
    }
    return render(request, 'contabilidad/confirmar_eliminar_comprobante.html', context) # Crea esta plantilla

def exportar_reporte_contable_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_contable.xlsx"'

    workbook = Workbook()

    # --- Hoja 1: Comprobantes y Detalles ---
    sheet_comprobantes = workbook.active
    sheet_comprobantes.title = "Comprobantes y Detalles"

    # Estilos
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Encabezados
    headers = [
        'Número Comprobante', 'Fecha Comprobante', 'Tipo Comprobante', 'Estado Comprobante', 
        'Descripción Comprobante', 'Total Débito (Comp.)', 'Total Crédito (Comp.)',
        'Tipo Doc. Fuente', 'ID Doc. Fuente', 'Fecha Creación (Comp.)',
        'Código Cuenta', 'Nombre Cuenta', 'Código SubCuenta', 'Nombre SubCuenta',
        'Monto Débito (Detalle)', 'Monto Crédito (Detalle)', 'Descripción Línea Detalle',
        'Número Factura Asociada', 'Tipo Factura', 'Estado Factura', 'Monto Total Factura', 
        'Nombre Cliente/Proveedor Factura', 'Fecha Factura'
    ]
    sheet_comprobantes.append(headers)
    for col_num, cell in enumerate(sheet_comprobantes[1]):
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        sheet_comprobantes.column_dimensions[chr(65 + col_num)].width = 20 # Ancho de columna por defecto

    comprobantes = ComprobanteMaestro.objects.all().order_by('-fecha_comprobante')

    for comp in comprobantes:
        detalles = comp.detalles.all()
        factura_asociada = getattr(comp, 'factura_asociada', None)

        if not detalles:
            row_data = [
                comp.numero_comprobante, str(comp.fecha_comprobante), comp.get_tipo_comprobante_display(), 
                comp.get_estado_display(), comp.descripcion, float(comp.total_debito), float(comp.total_credito),
                comp.tipo_documento_fuente, comp.id_documento_fuente, str(comp.fecha_creacion),
                '', '', '', '', 0.00, 0.00, '',
                factura_asociada.numero_factura if factura_asociada else '',
                factura_asociada.get_tipo_factura_display() if factura_asociada else '',
                factura_asociada.get_estado_display() if factura_asociada else '',
                float(factura_asociada.monto_total) if factura_asociada else 0.00,
                factura_asociada.nombre_cliente_proveedor if factura_asociada else '',
                str(factura_asociada.fecha_factura) if factura_asociada else ''
            ]
            sheet_comprobantes.append(row_data)
        else:
            for det in detalles:
                row_data = [
                    comp.numero_comprobante, str(comp.fecha_comprobante), comp.get_tipo_comprobante_display(), 
                    comp.get_estado_display(), comp.descripcion, float(comp.total_debito), float(comp.total_credito),
                    comp.tipo_documento_fuente, comp.id_documento_fuente, str(comp.fecha_creacion),
                    det.cuenta_mayor.codigo_cuenta, det.cuenta_mayor.nombre_cuenta,
                    det.sub_cuenta.codigo_sub_cuenta if det.sub_cuenta else '',
                    det.sub_cuenta.nombre_sub_cuenta if det.sub_cuenta else '',
                    float(det.monto_debito), float(det.monto_credito), det.descripcion_linea,
                    factura_asociada.numero_factura if factura_asociada else '',
                    factura_asociada.get_tipo_factura_display() if factura_asociada else '',
                    factura_asociada.get_estado_display() if factura_asociada else '',
                    float(factura_asociada.monto_total) if factura_asociada else 0.00,
                    factura_asociada.nombre_cliente_proveedor if factura_asociada else '',
                    str(factura_asociada.fecha_factura) if factura_asociada else ''
                ]
                sheet_comprobantes.append(row_data)

    # --- Hoja 2: Tipos de Cuenta ---
    sheet_tipos_cuenta = workbook.create_sheet(title="Tipos de Cuenta")
    sheet_tipos_cuenta.append(['ID', 'Nombre', 'Descripción', 'Naturaleza', 'Fecha Creación', 'Fecha Actualización'])
    for col_num, cell in enumerate(sheet_tipos_cuenta[1]):
        cell.font = header_font
        sheet_tipos_cuenta.column_dimensions[chr(65 + col_num)].width = 25

    for tipo in TipoCuenta.objects.all().order_by('nombre'):
        sheet_tipos_cuenta.append([
            tipo.id, tipo.nombre, tipo.descripcion, tipo.get_naturaleza_cuenta_display(), 
            str(tipo.fecha_creacion), str(tipo.fecha_actualizacion)
        ])

    # --- Hoja 3: Cuentas Mayor ---
    sheet_cuentas_mayor = workbook.create_sheet(title="Cuentas Mayor")
    sheet_cuentas_mayor.append([
        'ID', 'Código', 'Nombre', 'Tipo Cuenta', 'Activa', 'Permite Transacciones', 'Descripción', 
        'Fecha Creación', 'Fecha Actualización'
    ])
    for col_num, cell in enumerate(sheet_cuentas_mayor[1]):
        cell.font = header_font
        sheet_cuentas_mayor.column_dimensions[chr(65 + col_num)].width = 25

    for cuenta_m in CuentaMayor.objects.all().order_by('codigo_cuenta'):
        sheet_cuentas_mayor.append([
            cuenta_m.id, cuenta_m.codigo_cuenta, cuenta_m.nombre_cuenta, cuenta_m.tipo_cuenta.nombre,
            'Sí' if cuenta_m.esta_activa else 'No', 
            'Sí' if cuenta_m.permite_transacciones else 'No',
            cuenta_m.descripcion, str(cuenta_m.fecha_creacion), str(cuenta_m.fecha_actualizacion)
        ])

    # --- Hoja 4: Sub-Cuentas ---
    sheet_sub_cuentas = workbook.create_sheet(title="Sub-Cuentas")
    sheet_sub_cuentas.append([
        'ID', 'Código Sub-Cuenta', 'Nombre Sub-Cuenta', 'Cuenta Mayor (Código)', 'Cuenta Mayor (Nombre)', 
        'Activa', 'Descripción', 'Fecha Creación', 'Fecha Actualización'
    ])
    for col_num, cell in enumerate(sheet_sub_cuentas[1]):
        cell.font = header_font
        sheet_sub_cuentas.column_dimensions[chr(65 + col_num)].width = 25

    for sub_c in SubCuenta.objects.all().order_by('codigo_sub_cuenta'):
        sheet_sub_cuentas.append([
            sub_c.id, sub_c.codigo_sub_cuenta, sub_c.nombre_sub_cuenta, 
            sub_c.cuenta_mayor.codigo_cuenta, sub_c.cuenta_mayor.nombre_cuenta,
            'Sí' if sub_c.esta_activa else 'No', 
            sub_c.descripcion, str(sub_c.fecha_creacion), str(sub_c.fecha_actualizacion)
        ])

    workbook.save(response)
    return response
